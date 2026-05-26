"""Export shipment info and multi-row rate headers to Excel."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from customization_per_carrier import (
    apply_display_name_override,
    detect_carrier_key,
    get_shipment_columns,
    get_skip_rate_columns,
    get_standard_display_names,
    resolve_applies_if,
    resolve_rate_by,
)
from format_rates_layout import format_rates_workbook
from config import OUTPUT_DIR
from transform_ratebook import FRONT_COLUMN_ORDER, normalize_column_name

HEADER_ROWS = 6
BLOCK_ORIGIN = "Origin Charges"
BLOCK_MAIN = "Main freight Charges"
BLOCK_DEST = "Destination charges"
BLOCK_ORDER = [BLOCK_ORIGIN, BLOCK_MAIN, BLOCK_DEST]

WEIGHT_BRACKET_RE = re.compile(
    r"([\d][\d,]*)\s*-\s*([\d][\d,]*)\s*kg",
    re.IGNORECASE,
)

# Block detection from column name tokens (not per-cost mappings).
ORIGIN_TOKENS = ("pre carriage", "pre-carriage", "origin")
MAIN_TOKENS = ("main carriage", "pss")
DEST_TOKENS = ("on carriage", "destination")

@dataclass
class RateSubColumn:
    header: str
    source_column: str | None = None
    is_currency: bool = False
    min_max_label: str | None = None


@dataclass
class RateCost:
    block: str
    name: str
    applies_if: str
    rate_by: str
    sub_columns: list[RateSubColumn] = field(default_factory=list)


def parse_upper_weight_bound(column_name: str) -> int | None:
    match = WEIGHT_BRACKET_RE.search(normalize_column_name(column_name))
    if not match:
        return None
    return int(match.group(2).replace(",", ""))


def is_weight_bracket_column(name: str) -> bool:
    return parse_upper_weight_bound(name) is not None


def weight_bracket_header(name: str) -> str:
    upper = parse_upper_weight_bound(name)
    return f"<={upper}" if upper is not None else normalize_column_name(name)


def is_min_column(name: str) -> bool:
    norm = normalize_column_name(name)
    return "minimum charge" in norm or re.search(r"\bmin\b", norm)


def is_max_column(name: str) -> bool:
    norm = normalize_column_name(name)
    return re.search(r"\bmax\b", norm)


def is_per_kg_column(name: str) -> bool:
    return "per kg" in normalize_column_name(name)


def classify_block(column_name: str) -> str:
    norm = normalize_column_name(column_name)
    if any(token in norm for token in ORIGIN_TOKENS):
        return BLOCK_ORIGIN
    if any(token in norm for token in DEST_TOKENS):
        return BLOCK_DEST
    if any(token in norm for token in MAIN_TOKENS) or is_weight_bracket_column(column_name):
        return BLOCK_MAIN
    return BLOCK_MAIN


def cost_column_stem(column: str) -> str:
    text = normalize_column_name(column)
    text = WEIGHT_BRACKET_RE.sub(" ", text)
    text = re.sub(r"\bminimum charge\b", " ", text)
    text = re.sub(r"\b(min|max|per kg|in days)\b", " ", text)
    text = re.sub(r"\b(charge|fee)\b", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def longest_common_stem(stems: list[str]) -> str:
    if not stems:
        return ""
    common = stems[0]
    for stem in stems[1:]:
        limit = min(len(common), len(stem))
        idx = 0
        while idx < limit and common[idx] == stem[idx]:
            idx += 1
        common = common[:idx]
    return common.strip()


def infer_display_name(columns: list[str]) -> str:
    stems = [cost_column_stem(col) for col in columns if not is_weight_bracket_column(col)]
    if not stems:
        stems = [cost_column_stem(columns[0])]
    name = longest_common_stem(stems) or cost_column_stem(columns[0])
    if not name:
        name = normalize_column_name(columns[0])
    return name.title()


def infer_applies_if(columns: list[str]) -> str:
    norms = [normalize_column_name(col) for col in columns]
    if any("customs clearance" in norm for norm in norms):
        return ""
    if sum(is_weight_bracket_column(col) for col in columns) >= 2:
        return "Applies if: 1. Equipment Type equals 'LTL/Standard"
    if any(is_min_column(col) or is_per_kg_column(col) for col in columns):
        return "Applies if invoiced by Carrier"
    if any("documentation" in norm for norm in norms):
        return "Applies if invoiced by Carrier"
    return ""


def infer_rate_by(columns: list[str]) -> str:
    norms = [normalize_column_name(col) for col in columns]
    if any("customs clearance" in norm for norm in norms):
        if any(token in norm for token in ORIGIN_TOKENS for norm in norms):
            return "Customs/Exp"
        return "Customs/Imp"
    if any("documentation" in norm for norm in norms):
        return "per shipment"
    if any("transit time" in norm for norm in norms):
        return ""
    if any(
        is_per_kg_column(col) or is_weight_bracket_column(col) or is_min_column(col)
        for col in columns
    ):
        return "Weight/chargeable kg"
    return ""


def build_sub_columns(source_columns: list[str]) -> list[RateSubColumn]:
    has_weight_brackets = any(is_weight_bracket_column(col) for col in source_columns)
    sub_columns = [RateSubColumn(header="Currency", is_currency=True)]
    for source in source_columns:
        if is_min_column(source):
            if has_weight_brackets:
                sub_columns.append(
                    RateSubColumn(
                        header="p/unit",
                        source_column=source,
                        min_max_label="MIN",
                    )
                )
            else:
                sub_columns.append(
                    RateSubColumn(
                        header="Flat",
                        source_column=source,
                        min_max_label="MIN",
                    )
                )
        elif is_max_column(source):
            sub_columns.append(
                RateSubColumn(
                    header="p/unit" if has_weight_brackets else "Flat",
                    source_column=source,
                    min_max_label="MAX",
                )
            )
        elif is_weight_bracket_column(source):
            sub_columns.append(
                RateSubColumn(
                    header="p/unit",
                    source_column=source,
                    min_max_label=weight_bracket_header(source),
                )
            )
        elif is_per_kg_column(source):
            sub_columns.append(RateSubColumn(header="p/unit", source_column=source))
        else:
            sub_columns.append(RateSubColumn(header="Flat", source_column=source))
    return sub_columns


def rate_columns_from_df(df: pd.DataFrame, carrier_key: str) -> list[str]:
    shipment_names = {
        normalize_column_name(name) for name in get_shipment_columns(carrier_key)
    }
    skip_columns = get_skip_rate_columns(carrier_key)
    rate_columns: list[str] = []
    for column in df.columns:
        norm = normalize_column_name(column)
        if norm in shipment_names or norm in skip_columns:
            continue
        rate_columns.append(column)
    return rate_columns


def group_rate_columns(columns: list[str]) -> list[list[str]]:
    """Group adjacent rate columns into one cost using column-name patterns."""
    groups: list[list[str]] = []
    idx = 0
    while idx < len(columns):
        column = columns[idx]

        if is_min_column(column) or is_max_column(column):
            group = [column]
            idx += 1
            while idx < len(columns):
                next_col = columns[idx]
                if is_per_kg_column(next_col):
                    group.append(next_col)
                    idx += 1
                    break
                if is_weight_bracket_column(next_col):
                    group.append(next_col)
                    idx += 1
                    continue
                break
            groups.append(group)
            continue

        if is_weight_bracket_column(column):
            group = [column]
            idx += 1
            while idx < len(columns) and is_weight_bracket_column(columns[idx]):
                group.append(columns[idx])
                idx += 1
            groups.append(group)
            continue

        if is_per_kg_column(column):
            groups.append([column])
            idx += 1
            continue

        groups.append([column])
        idx += 1
    return groups


def is_zero_value(value) -> bool:
    if value is None or pd.isna(value):
        return True
    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return True
        try:
            return float(text.replace(",", "")) == 0.0
        except ValueError:
            return False
    try:
        return float(value) == 0.0
    except (TypeError, ValueError):
        return False


def cost_has_nonzero_values(df: pd.DataFrame, cost: RateCost) -> bool:
    for sub in cost.sub_columns:
        if sub.is_currency or not sub.source_column:
            continue
        if not df[sub.source_column].map(is_zero_value).all():
            return True
    return False


def remove_all_zero_cost_groups(costs: list[RateCost], df: pd.DataFrame) -> list[RateCost]:
    kept = [cost for cost in costs if cost_has_nonzero_values(df, cost)]
    removed = len(costs) - len(kept)
    if removed:
        print(f"Removed {removed} cost group(s) with only zero values (excluding Currency).")
    return kept


def build_rate_cost_groups(df: pd.DataFrame, carrier_key: str) -> list[RateCost]:
    columns = rate_columns_from_df(df, carrier_key)
    by_block: dict[str, list[RateCost]] = {block: [] for block in BLOCK_ORDER}

    for group_columns in group_rate_columns(columns):
        block = classify_block(group_columns[0])
        inferred_name = infer_display_name(group_columns)
        display_name = apply_display_name_override(
            inferred_name, group_columns, carrier_key
        )
        by_block[block].append(
            RateCost(
                block=block,
                name=display_name,
                applies_if=resolve_applies_if(
                    display_name,
                    group_columns,
                    infer_applies_if(group_columns),
                    carrier_key,
                ),
                rate_by=resolve_rate_by(
                    display_name,
                    group_columns,
                    infer_rate_by(group_columns),
                    carrier_key,
                ),
                sub_columns=build_sub_columns(group_columns),
            )
        )

    ordered: list[RateCost] = []
    for block in BLOCK_ORDER:
        ordered.extend(by_block[block])
    return remove_all_zero_cost_groups(ordered, df)


def get_currency_series(df: pd.DataFrame) -> pd.Series:
    if "Currency" in df.columns:
        return df["Currency"]
    return pd.Series("USD", index=df.index)


def write_merged_row(
    ws,
    row_idx: int,
    start_col: int,
    end_col: int,
    value: str,
) -> None:
    if end_col > start_col:
        ws.merge_cells(
            start_row=row_idx,
            end_row=row_idx,
            start_column=start_col,
            end_column=end_col,
        )
    ws.cell(row=row_idx, column=start_col, value=value)


def export_rates_layout_xlsx(
    df: pd.DataFrame,
    source_path: Path,
    sheet_name: str,
    carrier_key: str | None = None,
) -> Path:
    carrier_key = carrier_key or detect_carrier_key(df)
    print(f"Using carrier customization profile: {carrier_key}")
    shipment_columns = get_shipment_columns(carrier_key)
    standard_display_names = get_standard_display_names(carrier_key)
    rate_costs = build_rate_cost_groups(df, carrier_key)
    currency = get_currency_series(df)

    wb = Workbook()
    ws = wb.active
    ws.title = re.sub(r'[\\/*?:\[\]]', "_", sheet_name)[:31]

    shipment_width = len(shipment_columns)
    col_cursor = shipment_width + 1
    cost_spans: list[tuple[RateCost, int, int]] = []

    for cost in rate_costs:
        start = col_cursor
        col_cursor += len(cost.sub_columns)
        cost_spans.append((cost, start, col_cursor - 1))

    block_spans: dict[str, tuple[int, int]] = {}
    for block in BLOCK_ORDER:
        block_costs = [span for span in cost_spans if span[0].block == block]
        if block_costs:
            block_spans[block] = (block_costs[0][1], block_costs[-1][2])

    for row_idx in range(1, HEADER_ROWS + 1):
        for col_idx in range(1, shipment_width + 1):
            ws.cell(row=row_idx, column=col_idx, value=None)

    for block, (start_col, end_col) in block_spans.items():
        write_merged_row(ws, 1, start_col, end_col, block)

    for cost, start_col, end_col in cost_spans:
        write_merged_row(ws, 2, start_col, end_col, cost.name)
        write_merged_row(ws, 3, start_col, end_col, cost.applies_if)
        write_merged_row(ws, 4, start_col, end_col, cost.rate_by)

    for cost, start_col, _end_col in cost_spans:
        offset = start_col
        for sub in cost.sub_columns:
            ws.cell(row=HEADER_ROWS, column=offset, value=sub.header)
            if sub.min_max_label:
                ws.cell(row=HEADER_ROWS - 1, column=offset, value=sub.min_max_label)
            offset += 1

    for col_idx, header in enumerate(shipment_columns, start=1):
        ws.cell(row=HEADER_ROWS, column=col_idx, value=header)

    data_start_row = HEADER_ROWS + 1
    for row_offset, row_idx in enumerate(df.index):
        excel_row = data_start_row + row_offset
        for col_idx, column in enumerate(shipment_columns, start=1):
            ws.cell(row=excel_row, column=col_idx, value=df.at[row_idx, column])

        for cost, start_col, _end_col in cost_spans:
            col_pos = start_col
            for sub in cost.sub_columns:
                if sub.is_currency:
                    value = currency.at[row_idx]
                else:
                    value = df.at[row_idx, sub.source_column]
                ws.cell(row=excel_row, column=col_pos, value=value)
                col_pos += 1

    format_spans = [
        (
            cost.name,
            start_col,
            end_col,
            [(sub.header, sub.min_max_label) for sub in cost.sub_columns],
        )
        for cost, start_col, end_col in cost_spans
    ]
    format_rates_workbook(
        ws,
        shipment_columns,
        format_spans,
        data_start_row=data_start_row,
        data_row_count=len(df),
        standard_display_names=standard_display_names,
    )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    safe_sheet = re.sub(r'[<>:"/\\|?*]', "_", sheet_name)
    output_path = (
        OUTPUT_DIR / f"{source_path.stem}_{safe_sheet}_rates_layout.xlsx"
    )
    wb.save(output_path)
    return output_path
