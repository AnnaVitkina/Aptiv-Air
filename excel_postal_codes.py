"""Preserve postal / zip code leading zeros when reading Excel workbooks."""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

POSTAL_COLUMN_TOKENS = ("postal", "zip")


def is_postal_column(name: str) -> bool:
    norm = str(name).replace(" ", "").lower()
    return any(token in norm for token in POSTAL_COLUMN_TOKENS)


def _postal_pad_width(number_format: str) -> int | None:
    if not number_format or number_format in {"General", "@"}:
        return None
    zero_groups = re.findall(r"0+", number_format)
    if not zero_groups:
        return None
    return max(len(group) for group in zero_groups)


def cell_postal_value(cell) -> str | None:
    if cell.value is None:
        return None
    if cell.data_type == "s" or cell.number_format == "@":
        return str(cell.value).strip()
    if isinstance(cell.value, (int, float)):
        number = float(cell.value)
        if number.is_integer():
            integer = int(number)
            width = _postal_pad_width(cell.number_format or "")
            if width:
                return str(integer).zfill(width)
            return str(integer)
        return str(cell.value).strip()
    return str(cell.value).strip()


def patch_postal_code_columns(
    df: pd.DataFrame, xlsx_path: Path, sheet_name: str
) -> pd.DataFrame:
    postal_columns = [col for col in df.columns if is_postal_column(col)]
    if not postal_columns:
        return df

    wb = load_workbook(xlsx_path, read_only=True, data_only=False)
    try:
        ws = wb[sheet_name]
        header_cells = next(ws.iter_rows(min_row=1, max_row=1))
        header_map = {
            cell.value: cell.column - 1
            for cell in header_cells
            if cell.value is not None
        }

        result = df.copy()
        for column_name in postal_columns:
            result[column_name] = result[column_name].astype("object")

        for row_offset, row_idx in enumerate(df.index):
            excel_row = row_offset + 2
            for column_name in postal_columns:
                col_idx = header_map.get(column_name)
                if col_idx is None:
                    continue
                cell = ws.cell(row=excel_row, column=col_idx + 1)
                value = cell_postal_value(cell)
                if value is not None:
                    result.at[row_idx, column_name] = value
        return result
    finally:
        wb.close()


def postal_code_as_text(value) -> str | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text
